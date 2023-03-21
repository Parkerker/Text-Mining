import ExcelAccess as EA
import nltk
import os
import random
import numpy as np
import time
import re
import pickle
import csv

from math import log
from pandas import DataFrame as DF
from openpyxl import load_workbook
from nltk import word_tokenize
from nltk import sent_tokenize
from nltk.stem import SnowballStemmer 
from nltk.corpus import stopwords
from sklearn import feature_extraction
from sklearn.feature_extraction.text import TfidfVectorizer

docPath = "Source/SV/News train sets_5cate"
tfidfPath = "Source/PD/News"
catePath = "resultFile/Arranged/Label"
EX_STOP_WORD = set(stopwords.words("english"))|set(["https", "http","would","keep","like","also","could","one","two",
                                                    "without","may","want","even",'ever',"might","many","much","take",
                                                    "year","go","why","it"])

#'''語料庫'''
class NTLKCorpusSet:
    def __init__(self,doc=list(),feaSet=dict(),feaMode=str(),cateList=list(),allW=list(),trainFeaSet=list(),stemMode=False):
        print('A1')
        self.feaMode = feaMode
        #'''停用詞'''
        self.STOP_WORD = set(stopwords.words("english"))|set(["https", "http","would","keep","like","also","could","one","two",
                                                              "without","may","want","even",'ever',"might","many","much","take",
                                                              "year","go"])
        self.totalDocument = doc#'''所有文件'''
        self.featuerSet = feaSet#'''特徵字'''
        self.wordFreDict = nltk.FreqDist(doc)#'''文字出現頻率'''
        self.categoryList = cateList#'''類別清單'''
        self.allWords = allW#'''所有單字'''
        self.trainFeatuerSet = trainFeaSet#'''訓練功能集'''
        self.stemMode = stemMode#'''詞幹模式'''
        if self.stemMode:
            self.stemmer = SnowballStemmer("english")#'''刪除英文字尾'''
    #'''顯示文件狀態'''
    def showDocumentState(self):
        print('A2')
        print('Total Sheets:',len(self.totalDocument))#'''顯示總篇數'''
        print('All categorys:%s'%('\t'.join(category for category in self.categoryList)))#'''顯示所有類別'''
        print('Number of categorys:',len(self.categoryList))#'''顯示類別清單數量'''
        print('All words counts:%d'%(len(self.allWords)))#'''字數'''
        print('All words freq dict counts:%d'%(len(self.wordFreDict)))#顯示頻率字典數量
        print('Sheet of each category:')
        cateFreDict = self.getCategoryProportion()
        for cate,fre in cateFreDict.items():
            print(cate[:6],'\t',fre,'\t',round(fre/len(self.totalDocument)*100,3),'%')#顯示類別的比例
    #'''獲取各類別的數量'''
    def getCategoryProportion(self):
        print('A3')
        cateFreDict = {}
        tempCateSet = set()#'''set是一種集合'''
        for sheet in self.totalDocument:#'''取得各項的類別 然後進行加總'''
            if sheet[1] not in tempCateSet:
                tempCateSet.add(sheet[1])
                cateFreDict[sheet[1]] = 1
            else:
                cateFreDict[sheet[1]] += 1

        return cateFreDict
    #'''文檔設置'''
    def startSetDoc(self,path=docPath,mode='unarranged',conCol=2):
        print('A4')
        xlsxList = []
        xlsxList = EA.findDataFlienameList(path=path)
        

        if xlsxList:
            print("Find data file list :",xlsxList)
        else:
            print("Could not find data file.")
            return
        
        for xlsxFile in xlsxList:
            print(xlsxFile," is starting process.")
            if mode == 'unarranged':
                doc = EA.ArticleXlsx2NLTKFormat(xlsxFile,folderPath=path,contentCol=conCol)#'''取得文件'''
    
                if doc[0]['category'].lower() not in self.categoryList:
                    self.categoryList.append(doc[0]['category'].lower()) 

                for sheet in doc:
                    clrContent = list()
                    for token in sheet['content']:
                        invaildWord = re.search(r"[^a-zA-Z,.]+",token)#'''過濾多餘符號'''
                        if invaildWord is None :
                            if (token in self.STOP_WORD) or token == ',' or token == '.':
                                clrContent.append(token)
                                continue
                            token = token.lower()
                            if self.stemMode:
                                token = self.stemmer.stem(token)
                            clrContent.append(token)
                            self.allWords.append(token)

                    if clrContent:#'''輸入進totalDocument中'''
                        clrDoc = (clrContent,sheet['category'])
                        self.totalDocument.append(clrDoc)


                    
                print(xlsxFile,"was Done.")
                
            elif mode == 'arranged':
                for xlsxFile in xlsxList:
                    self.totalDocument = EA.arrangeArticleXlsx2NLTKFormat(xlsxFile)
                
                xlsxFile = EA.findDataFlienameList(catePath)
                wb = load_workbook(xlsxFile[0])#'''讀取檔案'''
                active_sheet=wb.active
                sheet = wb.worksheets[0]#'''工作表'''
                #para: min_row, max_row, min_column, max_column
                row_count = sheet.max_row
                for i in range(1, row_count):
                    category= active_sheet.cell(row=i, column=1).value.lower()
                    self.categoryList.append(category)

        self.wordFreDict = nltk.FreqDist(self.allWords)#'''計算單字出現頻率'''
        self.showDocumentState()
        #self.saveTotalDoc()
        random.shuffle(self.totalDocument)
        
    #'''設定特徵'''
    def setFeature(self,sheet):
        print('A5')
        features = {}
        #print(document)
        if self.feaMode == 'highFre':
            for w in sheet:
                features[w] = (w in self.featuerSet)

        elif self.feaMode == 'tfidf':
            for w in sheet:
                for cate,words in self.featuerSet.items():
                    if w in words:
                        features[w] = cate
        
        return features
    #'''尋找特徵字'''
    def findFeatureWords(self,featureLen=3000,tfidfFilePath=os.getcwd()+"/tfidfFile/categoryRevise"):
        print('A6')
        print("Feature Mode :",self.feaMode)
        if self.feaMode == 'tfidf':
            tfidfFileList = EA.findDataFlienameList(path=tfidfFilePath)
            print("Find xlsx file list :",tfidfFileList)
            for tfidfFile in tfidfFileList:
                
                wb = load_workbook(tfidfFilePath+ "/" + tfidfFile)
                active_sheet=wb.active
                sheet = wb.worksheets[0]
                print(tfidfFilePath+ "/" + tfidfFile)
                
                category= active_sheet.cell(row=1, column=1).value.lower()
                print('fff')
                if category not in self.categoryList:
                    continue
                #para: min_row, max_row, min_column, max_column
                feaNum = min(sheet.max_row,featureLen)
                wordsBag = set()
                wordCount = 0
                for i in range(2, sheet.max_row):
                    featureWord = active_sheet.cell(row=i, column=2).value
                    #'''前面的檢查是否是停用字後面的檢查是否為數字組成的'''
                    if (featureWord not in self.STOP_WORD) and (not str.isdigit(featureWord)):
                        if self.stemMode:
                            wordsBag.add(self.stemmer.stem(featureWord))
                        else:
                            wordsBag.add(featureWord)
                        wordCount += 1
                        if  wordCount >= feaNum:
                            break
                self.featuerSet[category] = wordsBag#'''更新特徵字'''

            

            for cate,words in self.featuerSet.items():
                print(cate,list(words)[:15]) #'''顯示類別和前15個'''

        elif self.feaMode == 'highFre':
            word_f = sorted(self.wordFreDict.items(),key=lambda d:d[1],reverse=True)[:featureLen]#'''排序單字'''
            self.featuerSet = set([key for (key,value) in word_f])
        
    #'''特徵集設置'''
    def setFeatureSet(self):
        print('A7')
        print("Start Setting Feature Set.")
        sTime = time.time()#'''當前時間'''
        self.trainFeatuerSet = [(self.setFeature(sheet), category) for (sheet, category) in self.totalDocument]
        eTime = time.time()
        print("Feature Set was set, Cost time :",eTime-sTime)
        print("Feature Set length : ",len(self.trainFeatuerSet))

#'''分類器''''
class NTLKClassifier:
    def __init__(self,clsifier = None):
        print('B1')
        self.clsfier = clsifier
        
    #'''訓練'''
    def startTrain(self,dataset,splitPercent=0.75):#'''splitPercent調整百分比'''
        print('B2')
        splitPoint = int(len(dataset.trainFeatuerSet)*splitPercent)
        print("Start train. Split point = ",splitPoint) 
        training_set = dataset.trainFeatuerSet[:splitPoint]#'''取出前0.75%資料'''
        sTime = time.time()
        classifier = nltk.NaiveBayesClassifier.train(training_set)#'''訓練貝氏分類器'''
        eTime = time.time()

        self.clsfier = classifier
        print("Train Over.  Cost time :",eTime-sTime)
    #'''精準度測試'''
    def accuracyTest(self,dataset,splitPercent=0.75):
        print('B3')
        trainAccs = []
        testAccs = []
        splitPoint = int(len(dataset.trainFeatuerSet)*splitPercent)
        training_set = dataset.trainFeatuerSet[:splitPoint]#'''前0.75%資料'''
        testing_set = dataset.trainFeatuerSet[splitPoint:]#'''後0.25%資料'''
        print("Accuracy test...")
        sTime = time.time()
        train_acc = (nltk.classify.accuracy(self.clsfier, training_set))*100#'''顯示其精準度'''
        test_acc = (nltk.classify.accuracy(self.clsfier, testing_set))*100
        trainAccs.append(train_acc)
        testAccs.append(test_acc)
        eTime = time.time()

        print("Cost time :",eTime-sTime)
        print("Total train accuracy percent:",trainAccs)
        print("Total test accuracy percent:",testAccs)
        print("Mean train accuracy percent:", np.mean(trainAccs))#'''取平均值'''
        print("Mean test accuracy percent:", np.mean(testAccs))
        #得到似然比，檢測那些是有用的特徵
        self.clsfier.show_most_informative_features(50)
    #部分精度測試
    def partAccuracyTest(self,dataset):
        print('B4')
        cateDict = dict.fromkeys(dataset.categoryList)#創建資料類別字典
        cateList = []
        numOfShtList = []
        baseAccList = []
        accList = []
        
        for cate,_ in cateDict.items():
            cateDict[cate] = []#'''將內容清空'''
            
        for sheet in dataset.trainFeatuerSet:
            cateDict[sheet[1]].append(sheet)#'''將文章塞到對應類別中'''

          
        cateAccDict = dict.fromkeys(dataset.categoryList)#創建類別計數字典

        for cate,_ in cateAccDict.items():
            cateAccDict[cate] = 0#全部初始化

        for cate,docSet in cateDict.items():#
            cateList.append(cate)
            numOfShtList.append(len(docSet))
            baseAccList.append(round(len(docSet)/len(dataset.totalDocument),3))
            accList.append(round(nltk.classify.accuracy(self.clsfier,docSet),3))

            print("Category :",cate)
            print("Number of sheets :",len(docSet))
            print("Base Accuracy :",round(len(docSet)/len(dataset.totalDocument)*100,3),'%')
            cateAcc = round(nltk.classify.accuracy(self.clsfier,docSet)*100,3)
            
            print("Accuracy :",cateAcc,"%\n")

        DFDict = {"Category": cateList,#類別 張數 基本精確度(全部類別的) 部件精確度
                  "Sheets Count": numOfShtList,
                  "Base Accuracy": baseAccList,
                  "Accuracy":accList
                  }
        cateDF = DF(DFDict)#將字典轉成資料型態
        print(cateDF)

    #'''分類器測試'''
    def classifieTest(self,dataset):
        print('B5')
        article = input("Enter your article:")#輸入文章
        feaArticle = dataset.setFeature(word_tokenize(article))#斷句
        print(feaArticle)
        print("Your article probably about :",self.clsfier.classify(feaArticle))#測試
        print("Probility of Article label :")
        for cate,prob in self.clsfier.prob_classify(feaArticle)._prob_dict.items():#判斷為各類別可能性
            print(cate,':',round(prob,5))
#'''PMI計算'''
class NLTKPMIcomputer:
    def __init__(self,pmiArray = dict()):
        print('C1')
        self.pmiArray = pmiArray
    #'''單詞包含索引集'''
    def wordContainIndexSet(self,word,sentsList):
        print('C2')
        containIndexSet = set()
        for i in range(len(sentsList)):
            if word in sentsList[i]:
                containIndexSet.add(i)
        return containIndexSet
    #'''句子包含索引集'''
    def wordTogetherIndexSet(self,cw1set,cw2set):
        print('C3')
        togetherIndexSet = cw1set&cw2set
        return togetherIndexSet
    #'''字出現的概率'''
    def wordProbility(self,word,sentsList):
        print('C4')
        count = 1
        for sents in sentsList:
            if word in sents:
                count += 1 
        return count/float(len(sentsList))
        
    #'''字同時出現的概率'''
    def wordTogetherProbility(self,word1,word2,sentsList):
        print('C5')
        togetherCount = 1
        for sents in sentsList:
            if (word1 in sents) and (word2 in sents) :
                togetherCount += 1
        return togetherCount/float(len(sentsList))

   
        
    #'''保存PMI陣列'''
    def savePMIarray(self,path='resultFile',name='PMIDict.csv'):
        print('C6')
        for cate,PMIdict in self.pmiArray.items():
            with open(path+"/"+cate+name, 'w', newline='') as csvfile:#開啟檔案並處理,此用法處理完會自動關閉檔案
                writer = csv.writer(csvfile)#寫csv檔案內容
                firstRow = list([cate])
                firstRow += PMIdict.keys()#返回字典內的鍵值
                for rowCount in range(1,len(PMIdict)+2):
                    if rowCount == 1:
                        writer.writerow(firstRow)
                    else:
                        word1 = firstRow[rowCount-1]
                        PMIList = list([word1])
                        for word2Count in range(1,len(firstRow)):
                            PMIList.append(PMIdict[word1][firstRow[word2Count]])#將這篇的文字兩兩對應

                        writer.writerow(PMIList)#寫入
            print(cate+name,"was saved.")
        
    #'''PMI計算'''
    def sheetPMI(self,sentsList,word1,word2,logMode=True):
        print('C7')
        if len(sentsList) == 0:
            return 0 
        sentsListLen = len(sentsList)
        cw1IndexSet = self.wordContainIndexSet(word1,sentsList)
        w1Probility = (len(cw1IndexSet))/sentsListLen#單字1出現的機率
        cw2IndexSet = self.wordContainIndexSet(word2,sentsList)
        w2Probility = (len(cw2IndexSet))/sentsListLen#單字2出現的機率
        togetherIndexSet =  self.wordTogetherIndexSet(cw1IndexSet,cw2IndexSet)
        togetherProbility = (len(togetherIndexSet))/sentsListLen#單字1和2同時出現的機率
        if w1Probility == 0 or w2Probility == 0 or togetherProbility == 0 :#PMI計算
            return 0
        PMI = togetherProbility/(w1Probility*w2Probility)
        if logMode:
            PMI = log(togetherProbility/(w1Probility*w2Probility),2)
        return PMI

    #'''設置PMI陣列'''
    def setPMIArray(self,dataset,wordlist,stem=True):
        print('C8')
        documentSet = dataset.totalDocument
        if stem :
            stemmer = SnowballStemmer("english")#刪除英文字尾
            wordlist = [stemmer.stem(word) for word in wordlist]
        #--PMIDict : PMI[category][word1][word2] = PMI value--#
        PMIDict = dict.fromkeys(dataset.categoryList)#將各類別先創建好
        print("Word List Length :",len(wordlist))#有多少字
        print("Initial PMI Dictionary..")
        for cate,_1 in PMIDict.items():
            PMIDict[cate] = dict.fromkeys(wordlist)#將每個字塞到類別中
            for word1,_2 in PMIDict[cate].items():
                PMIDict[cate][word1] = dict.fromkeys(wordlist)#將字兩兩相對
                for word2,_3 in PMIDict[cate][word1].items():
                    PMIDict[cate][word1][word2] = 0#將其初始化
        print("PMI Dictionary was initialed")

        print("Start computing PMI..")
        print("Number of sheets : " ,len(documentSet))#顯示篇數
        sheetCount = 0

        cateNumDict = dict.fromkeys(dataset.categoryList)#設置類別
        for cate,value in cateNumDict.items():
            cateNumDict[cate] = 0#初始化

        for doc in documentSet:
            category = doc[1]#類別
            content = doc[0]#內容
            if stem :
                content = [stemmer.stem(token) for token in content]
            cnctDoc = " ".join(content)
            sentsList = sent_tokenize(cnctDoc)#分割句子
            sentsListLen = len(sentsList)#計算數量
            cateNumDict[category] += 1#該類別數量加一
            #PMI計算且紀錄
            for word1Count in range(0,len(wordlist)):
                word1 = wordlist[word1Count]
                cw1IndexSet = self.wordContainIndexSet(word1,sentsList)
                w1Probility = (len(cw1IndexSet)+1)/sentsListLen
                for word2Count in range(word1Count+1,len(wordlist)):
                    word2 = wordlist[word2Count]
                    cw2IndexSet = self.wordContainIndexSet(word2,sentsList)
                    w2Probility = (len(cw2IndexSet)+1)/sentsListLen
                    togetherIndexSet =  self.wordTogetherIndexSet(cw1IndexSet,cw2IndexSet)
                    togetherProbility = len(togetherIndexSet)/sentsListLen
                    tempPMI = togetherProbility/(w1Probility*w2Probility)
                    PMIDict[category][word1][word2] += tempPMI

                    #--PMI(w1,w2) = PMI(w2, w1)--#
                    PMIDict[category][word2][word1] = PMIDict[category][word1][word2]

            sheetCount += 1 
            if sheetCount % 5000 == 0:
                print("Number of Sentence :",len(sentsList))#句數
                print(sheetCount," done . ")#篇數
        
        print("First PMI Done. Start Disjust..")
        #--disjust PMI value--#
        for category in dataset.categoryList:
            for word1 in wordlist:
                for word2 in wordlist:
                    if PMIDict[category][word1][word2] > 0: 
                        PMIDict[category][word1][word2] = log(PMIDict[category][word1][word2]/cateNumDict[category],2)#正式的PMI數值


        print("Disjust done.")
        self.pmiArray = PMIDict

#'''TFIDF計算器'''
class NLTKTFIDFComputer:
    def __init__(self,tfidfArr = list()):
        print('D1')
        self.TFIDFArray = tfidfArr
        self.stopWord = set(stopwords.words("english"))|set(["https", "http","would","keep","like","also","could","one","two",
                                                        "without","may","want","even",'ever',"might","many","much","take",
                                                        "year","go","as"])
    #'''TFIDF計算'''
    def TFIDF_Compute(self,fileName, text, cate):
        print('D2')
        destinationPath = str(os.getcwd())+fileName+'.pkl'
        file = open(destinationPath, 'rb')
        obj = pickle.load(file)
        file.close()

        split_word = ''
        for word in nltk.word_tokenize(text):#進行斷句
            word = word.lower()#轉成全小寫
            if str(word) == ',' or str(word) == '.' or (word in self.stopWord ):
                continue
            split_word += str(word) + ' '

        corpus = []
        corpus.append(split_word)
        corpus.append(str(obj[0]))

        vectorizer = TfidfVectorizer()#轉成tf-idf的特徵矩陣
        tfidf = vectorizer.fit_transform(corpus)#資料預處理
        #print(tfidf.shape, '(Files, Words)')  
        wordfeatures = vectorizer.get_feature_names()#特徵提取

        HF_words = []
        for i in range(len(wordfeatures)):
            if tfidf[0, i] > 0:
                temp = [wordfeatures[i], tfidf[0, i]]
                HF_words.append(temp)

        HF_words.sort(key=lambda s: s[1])
        HF_words.reverse()

        return HF_words




def main():
    print("main")
    feaMode = "tfidf" #'''使用TFIDF'''
    feaLen = 300
    contentCol = 2
    print("Full Process.")
    print("Feature Mode:",feaMode)
    print("Number of Features:",feaLen)
    print("Xlsx file content column:",contentCol)
    sTime = time.time()
    DataSet = NTLKCorpusSet(feaMode=feaMode,stemMode=False)
    DataSet.startSetDoc(conCol=contentCol)
    DataSet.findFeatureWords(featureLen=feaLen,tfidfFilePath=tfidfPath)
    DataSet.setFeatureSet()
    '''
    dataSetFile = open("dataSet"+time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())+".pickle","wb")
    pickle.dump(DataSet,dataSetFile)
    dataSetFile.close()
    
    Classifier = NTLKClassifier()

    Classifier.startTrain(DataSet)
    #Classifier.accuracyTest(DataSet)
    #Classifier.partAccuracyTest(DataSet)

    classifierFile = open("classifier"+time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())+".pickle","wb")
    pickle.dump(Classifier,classifierFile)
    classifierFile.close()

    eTime = time.time()
    print("Done. Cost time :",eTime-sTime)
    #self.classifieTest()
    '''
    '''
    feaWordList = list()
    for cate in DataSet.categoryList:
        feaWordList += list(DataSet.featuerSet[cate])[:10]
    
    PMIcomputer = NLTKPMIcomputer()
    PMIcomputer.setPMIArray(DataSet,feaWordList,stem=DataSet.stemMode)

    PMIcomputer.savePMIarray()
    '''


if __name__ == "__main__":
    main()