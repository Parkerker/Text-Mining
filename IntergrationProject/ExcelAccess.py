import os
import threading
import re
from openpyxl import load_workbook
from nltk import word_tokenize


currentPath = os.getcwd()
testFolder = "testFile/News train sets_8cate"
arrangedFolder = "resultFile/Arranged/Document"

class MyThread(threading.Thread):
    def __init__(self, target=None, args=(), **kwargs):
        super(MyThread, self).__init__()
        self._target = target
        self._args = args
        self._kwargs = kwargs

    def run(self):
        if self._target == None:
            return
        self.__result__ = self._target(*self._args, **self._kwargs)

    def get_result(self):
        self.join()       
        return self.__result__

def main():
    pass

def findDataFlienameList(path=currentPath+"/"+testFolder,subname='xlsx'):

    print(path)
    if not os.path.isabs(path):
        path = os.path.abspath(path)
    if os.path.exists(path):
        testFileList = os.listdir(path)
    else :
        print("Path dose not exist.")
        return
    
    xlsxList = list()
    for f in testFileList:
        if f.endswith(subname):
            xlsxList.append(f) 
    return xlsxList

def setArticleCateDict(article,category):
    TkedArtic = None
    categoryDict = dict()

    try:
        TkedArtic = word_tokenize(article)
    except TypeError as TE:
        pass
        
    if TkedArtic and TkedArtic != 'null':
        categoryDict['content'] = TkedArtic
        categoryDict['category'] = category
        return categoryDict


def ArticleXlsx2NLTKFormat(xlsxFile,folderPath=testFolder,contentCol=2):
    articleCateDictList = list()
    
    wb = load_workbook(currentPath + "/" +folderPath+ "/" + xlsxFile)
    active_sheet=wb.active 
    sheet = wb.worksheets[0]


    category= active_sheet.cell(row=1, column=1).value.lower()

    #para: min_row, max_row, min_column, max_column
    row_count = sheet.max_row
    for i in range(2, row_count):
        content = active_sheet.cell(row=i, column=contentCol).value
        
        articleCateDict = setArticleCateDict(content,category)
        if articleCateDict:
            articleCateDictList.append(articleCateDict)

        
    #print(tweetsCateDictList)
    return articleCateDictList    

def arrangeArticleXlsx2NLTKFormat(xlsxFile):
    articleCateList = list()
    
    wb = load_workbook(currentPath + "/" +arrangedFolder+ "/" + xlsxFile)
    active_sheet=wb.active
    sheet = wb.worksheets[0]

    #para: min_row, max_row, min_column, max_column
    row_count = sheet.max_row
    for i in range(2, row_count):
        category= active_sheet.cell(row=i, column=3).value.lower()
        content = active_sheet.cell(row=i, column=2).value
        
        articleCateDict = setArticleCateDict(content,category)
        if articleCateDict:
            articleCateList.append((articleCateDict['content'],articleCateDict['category']))

    #print(tweetsCateDictList)
    return articleCateList

def RetuenSingleXlsxDataFrame(Subpath):
    destinationPath=str(currentPath)+Subpath
    wb=load_workbook(destinationPath)
    active_sheet=wb.active
    sheet=wb.worksheets[0]
    DataFrame=[]
    DataFrame.append(sheet.max_row)
    DataFrame.append([])
    DataFrame.append([])
    for i in range(2, sheet.max_row):
        content = active_sheet.cell(row=i, column=2).value
        DataFrame[1].append(content)
        content = active_sheet.cell(row=i, column=3).value
        DataFrame[2].append(content)
    return DataFrame   


if __name__=="__main__":
    main()